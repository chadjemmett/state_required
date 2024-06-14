from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module="openpyxl")
html = open("state_req.html", "w")

links = []
# dict_data = {} 

doc = open("index.html", "r")
soup = bs(doc, "html.parser")

wb = load_workbook(filename='test_sheet.xlsx')
ws = wb['sheet']

for i in ws:
    dict_data = {}
    if i[0].offset(1, 1).value and i[0].value:
        dict_data['link_copy'] = i[0].value
        dict_data['link'] = i[3].value
        # print("children", i[0].value)
        cur_cell = i[0].offset(1, 1)
        dict_data['sub_items'] = []

        while cur_cell.offset(1, 0).value:
            # print("child\t",cur_cell.value)
            dict_data['sub_items'].append({'link_copy': cur_cell.value, 'link': cur_cell.offset(0, 2).value})
            cur_cell = cur_cell.offset(1, 0)
        dict_data['sub_items'].append({'link_copy': cur_cell.value, 'link': cur_cell.offset(0, 2).value})
        # print("child\t", cur_cell.value)
    else: 
        if i[0].value:
            dict_data['link'] = i[3].value
            dict_data['link_copy'] = i[0].value
# print("no children", i[0].value)

    links.append(dict_data)
            
        # cur_cell = i[0].offset(1, 1)
        # while cur_cell.offset(1, 0).value:
        #     cur_cell = cur_cell.offset(1, 0)
        #     print(i[0].value, "\n\t", cur_cell.value
# cur_cell = ws.cell(31, 1)
# while cur_cell.offset(1, 0).value:
#     print(cur_cell.value)
#     cur_cell = cur_cell.offset(1, 0)

# print(cur_cell.value)
for i in [x for x in links if len(x)]:
    print(i)
html.close()
